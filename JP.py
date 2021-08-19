from tkinter import * 
import tkinter as tk
import time
import random
import os.path
import xlrd
import xlsxwriter
import openpyxl
import logging as log

def fermer():
    global fenetre
    fenetre.destroy()

def Easy():
    niveau="easy"
    play(niveau)

def Normal():
    niveau="normal"
    play(niveau)

def Hard():
    niveau="hard"
    play(niveau)

def close(wnd):
    wnd.destroy()

def Classement():

    def EasyC():
        path = 'classement.xlsx'
        index1=0 # variable for take the index of the doc 1
        try :
            wb = xlrd.open_workbook(path) # To open Workbook
        except FileNotFoundError as e:
            def warning():
                # be treated as a new window
                warning = Toplevel(fenetre)
                
                # set the title 
                warning.title(" warning ")
                
                # sets the dimensions
                warning.geometry("850x125")
                
                # in the Window
                label = Label(warning, text=" You need to play for watch the scoreboard !", font=("Courier", 25))
                label.pack()
                btn_0 = Button(warning, text="Exit", fg='white',bg='#619ED6', height = 5, width = 8, command=warning.destroy)
                btn_0.pack(padx=5, pady=20)  
                #newWindow.iconbitmap('Back Up\\icon\\info.ico')
            warning()

        sheet = wb.sheet_by_index(index1) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        # copying the cell values from source 
        # excel file to destination excel file
        top1=sheet.cell_value(1, 1)
        top2=sheet.cell_value(1, 1)
        top3=sheet.cell_value(1, 1)
        top1N=sheet.cell_value(1, 0)
        top2N=" "
        top3N=" "
        for i in range (1, mr):
            wr1=sheet.cell_value(i, 1)
            if wr1 < top1 :
                top3=top2
                top2=top1
                top1=sheet.cell_value(i, 1)
                top1N=sheet.cell_value(i, 0)
            elif wr1 < top2 :
                top3=top2
                top2=sheet.cell_value(i, 1)
                top2N=sheet.cell_value(i, 0)
            elif wr1 < top3 :
                top3=sheet.cell_value(i, 1)
                top3N=sheet.cell_value(i, 0)
        lst=[["Classement","Name","Score"]]
        lst.append(["1 :",top1N,top1])
        lst.append(["2 :",top2N,top2])
        lst.append(["3 :",top3N,top3])

        n = 4
        master = Tk()
        lbox = Listbox(
            master,
            width=25,
            height=n,
            font="Verdana 20 bold",
            selectbackground="blue")
        lbox.pack(padx=20, pady=20,side=LEFT)
            
        for item in lst:
            lbox.insert(END, item)
            
        lbox.focus_set()
        pos = 1
        lbox.activate(pos)
        lbox.selection_set(pos)
            
        for i in range(0, len(lst), 2):
            lbox.itemconfigure(i, background='#f0f0ff')
        for i in range(1, len(lst), 2):
            lbox.itemconfigure(i, background='#fff')
        
        master.title("ScoreBoard")    

    def NormalC():
        path = 'classement.xlsx'
        index1=1 # variable for take the index of the doc 1
        try :
            wb = xlrd.open_workbook(path) # To open Workbook
        except FileNotFoundError as e:
            def warning():
                # be treated as a new window
                warning = Toplevel(fenetre)
                
                # set the title 
                warning.title(" warning ")
                
                # sets the dimensions
                warning.geometry("850x125")
                
                # in the Window
                label = Label(warning, text=" You need to play for watch the scoreboard !", font=("Courier", 25))
                label.pack()
                btn_0 = Button(warning, text="Exit", fg='white',bg='#619ED6', height = 5, width = 8, command=warning.destroy)
                btn_0.pack(padx=5, pady=20)  
                #newWindow.iconbitmap('Back Up\\icon\\info.ico')
            warning()

        sheet = wb.sheet_by_index(index1) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        # copying the cell values from source 
        # excel file to destination excel file
        top1=sheet.cell_value(1, 1)
        top2=sheet.cell_value(1, 1)
        top3=sheet.cell_value(1, 1)
        top1N=sheet.cell_value(1, 0)
        top2N=" "
        top3N=" "
        for i in range (1, mr):
            wr1=sheet.cell_value(i, 1)
            if wr1 < top1 :
                top3=top2
                top2=top1
                top1=sheet.cell_value(i, 1)
                top1N=sheet.cell_value(i, 0)
            elif wr1 < top2 :
                top3=top2
                top2=sheet.cell_value(i, 1)
                top2N=sheet.cell_value(i, 0)
            elif wr1 < top3 :
                top3=sheet.cell_value(i, 1)
                top3N=sheet.cell_value(i, 0)
        lst=[["Classement","Name","Score"]]
        lst.append(["1 :",top1N,top1])
        lst.append(["2 :",top2N,top2])
        lst.append(["3 :",top3N,top3])

        n = 4
        master = Tk()
        lbox = Listbox(
            master,
            width=25,
            height=n,
            font="Verdana 20 bold",
            selectbackground="blue")
        lbox.pack(padx=20, pady=20,side=LEFT)
            
        for item in lst:
            lbox.insert(END, item)
            
        lbox.focus_set()
        pos = 1
        lbox.activate(pos)
        lbox.selection_set(pos)
            
        for i in range(0, len(lst), 2):
            lbox.itemconfigure(i, background='#f0f0ff')
        for i in range(1, len(lst), 2):
            lbox.itemconfigure(i, background='#fff')
        
        master.title("ScoreBoard")    

    def HardC():
        path = 'classement.xlsx'
        index1=2 # variable for take the index of the doc 1
        try :
            wb = xlrd.open_workbook(path) # To open Workbook
        except FileNotFoundError as e:
            def warning():
                # be treated as a new window
                warning = Toplevel(fenetre)
                
                # set the title 
                warning.title(" warning ")
                
                # sets the dimensions
                warning.geometry("850x125")
                
                # in the Window
                label = Label(warning, text=" You need to play for watch the scoreboard !", font=("Courier", 25))
                label.pack()
                btn_0 = Button(warning, text="Exit", fg='white',bg='#619ED6', height = 5, width = 8, command=warning.destroy)
                btn_0.pack(padx=5, pady=20)  
                #newWindow.iconbitmap('Back Up\\icon\\info.ico')
            warning()

        sheet = wb.sheet_by_index(index1) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        # copying the cell values from source 
        # excel file to destination excel file
        top1=sheet.cell_value(1, 1)
        top2=sheet.cell_value(1, 1)
        top3=sheet.cell_value(1, 1)
        top1N=sheet.cell_value(1, 0)
        top2N=" "
        top3N=" "
        for i in range (1, mr):
            wr1=sheet.cell_value(i, 1)
            if wr1 < top1 :
                top3=top2
                top2=top1
                top1=sheet.cell_value(i, 1)
                top1N=sheet.cell_value(i, 0)
            elif wr1 < top2 :
                top3=top2
                top2=sheet.cell_value(i, 1)
                top2N=sheet.cell_value(i, 0)
            elif wr1 < top3 :
                top3=sheet.cell_value(i, 1)
                top3N=sheet.cell_value(i, 0)
        lst=[["Classement","Name","Score"]]
        lst.append(["1 :",top1N,top1])
        lst.append(["2 :",top2N,top2])
        lst.append(["3 :",top3N,top3])

        n = 4
        master = Tk()
        lbox = Listbox(
            master,
            width=25,
            height=n,
            font="Verdana 20 bold",
            selectbackground="blue")
        lbox.pack(padx=20, pady=20,side=LEFT)
            
        for item in lst:
            lbox.insert(END, item)
            
        lbox.focus_set()
        pos = 1
        lbox.activate(pos)
        lbox.selection_set(pos)
            
        for i in range(0, len(lst), 2):
            lbox.itemconfigure(i, background='#f0f0ff')
        for i in range(1, len(lst), 2):
            lbox.itemconfigure(i, background='#fff')
        
        master.title("ScoreBoard")    

    # be treated as a new window
    Classement = Toplevel(fenetre)
      
    # set the title 
    Classement.title("Classement !")
      
    # sets the dimensions
    Classement.geometry("525x150")
    
    btn_2 = Button(Classement, text="Hard", fg='white',bg='#FF968A', height = 20, width = 20, command=HardC)
    btn_2.pack(side=RIGHT, padx=5, pady=5) 
    btn_1 = Button(Classement, text="Normal", fg='white',bg='#97C1A9', height = 20, width = 20, command=NormalC)
    btn_1.pack(side=RIGHT, padx=5, pady=5)  
    btn_0 = Button(Classement, text="Easy", fg='white',bg='#8FCACA', height = 20, width = 20, command=EasyC)
    btn_0.pack(side=RIGHT, padx=5, pady=5)  
    
    btn_5 = Button(Classement, text="Exit", fg ='white',bg='black', height = 5, width = 5,command=Classement.destroy)
    btn_5.pack(side=LEFT, padx=5, pady=5)

def enregisterE(name,temps):

    path = 'classement.xlsx'
    if not os.path.exists(path) :
        workbook = xlsxwriter.Workbook(path) # Create a XLSX 
        worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Normal") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        workbook.close()

    def Copy(path, worksheet, index):
        wb = xlrd.open_workbook(path) # To open Workbook
        sheet = wb.sheet_by_index(index) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        for i in range (0, mr):
            for j in range (0, mc):
                # reading cell value from source excel file
                wr1=sheet.cell_value(i, j)
                # writing the read value to destination excel file
                worksheet.write(i, j, wr1) 
    
    workbook = xlsxwriter.Workbook(path) # Create a XLSX 
    worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
    Copy(path, worksheet,0)
    wb = xlrd.open_workbook(path) # To open Workbook
    sheet = wb.sheet_by_index(0) # variable for read the index of the doc1
    mr=sheet.nrows # number of the max row in the doc1
    mc=sheet.ncols # number of the max column in the doc1
    worksheet.write(mr, 0, name)
    worksheet.write(mr, 1, temps) 
    worksheet = workbook.add_worksheet("Nornal") # Tittle for the sheet
    Copy(path, worksheet,1)
    worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
    Copy(path, worksheet,2)
    workbook.close()

def enregisterN(name,temps):

    path = 'classement.xlsx'
    if not os.path.exists(path) :
        workbook = xlsxwriter.Workbook(path) # Create a XLSX 
        worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Normal") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        workbook.close()

    def Copy(path, worksheet, index):
        wb = xlrd.open_workbook(path) # To open Workbook
        sheet = wb.sheet_by_index(index) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        for i in range (0, mr):
            for j in range (0, mc):
                # reading cell value from source excel file
                wr1=sheet.cell_value(i, j)
                # writing the read value to destination excel file
                worksheet.write(i, j, wr1) 
    
    workbook = xlsxwriter.Workbook(path) # Create a XLSX 
    worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
    Copy(path, worksheet,0)
    worksheet = workbook.add_worksheet("Nornal") # Tittle for the sheet
    Copy(path, worksheet,1)
    wb = xlrd.open_workbook(path) # To open Workbook
    sheet = wb.sheet_by_index(1) # variable for read the index of the doc1
    mr=sheet.nrows # number of the max row in the doc1
    mc=sheet.ncols # number of the max column in the doc1
    worksheet.write(mr, 0, name)
    worksheet.write(mr, 1, temps) 
    worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
    Copy(path, worksheet,2)
    workbook.close()

def enregisterH(name,temps):

    path = 'classement.xlsx'
    if not os.path.exists(path) :
        workbook = xlsxwriter.Workbook(path) # Create a XLSX 
        worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Normal") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
        worksheet.write(0, 0, "Pseudo :")
        worksheet.write(0, 1, "Score :")
        workbook.close()

    def Copy(path, worksheet, index):
        wb = xlrd.open_workbook(path) # To open Workbook
        sheet = wb.sheet_by_index(index) # variable for read the index of the doc1
        mr=sheet.nrows # number of the max row in the doc1
        mc=sheet.ncols # number of the max column in the doc1
        for i in range (0, mr):
            for j in range (0, mc):
                # reading cell value from source excel file
                wr1=sheet.cell_value(i, j)
                # writing the read value to destination excel file
                worksheet.write(i, j, wr1) 
    
    workbook = xlsxwriter.Workbook(path) # Create a XLSX 
    worksheet = workbook.add_worksheet("Easy") # Tittle for the sheet
    Copy(path, worksheet,0)
    worksheet = workbook.add_worksheet("Nornal") # Tittle for the sheet
    Copy(path, worksheet,1)
    worksheet = workbook.add_worksheet("Hard") # Tittle for the sheet
    Copy(path, worksheet,2)
    wb = xlrd.open_workbook(path) # To open Workbook
    sheet = wb.sheet_by_index(2) # variable for read the index of the doc1
    mr=sheet.nrows # number of the max row in the doc1
    mc=sheet.ncols # number of the max column in the doc1
    worksheet.write(mr, 0, name)
    worksheet.write(mr, 1, temps) 
    workbook.close()

def afficher(name,temps):

    # be treated as a new window
    GG = Toplevel(fenetre)
    # set the title 
    GG.title("Well done")
        
    # sets the dimensions
    GG.geometry("600x400")
        
    # in the Window
    label = Label(GG, text="Bravo"+" "+name+ "\n", font=("Courier", 25))
    label.pack()
    label = Label(GG, text="Tu as fini en :"+" "+str(temps)+" secondes", font=("Courier", 15))
    label.pack()
    label = Label(GG, text="\n" + "Menu ?", font=("Courier", 25))
    label.pack()
    btn_0 = Button(GG, text="Yes", fg='white',bg='#619ED6', height = 5, width = 5, command=lambda: [close(GG)])
    btn_0.pack(padx=5, pady=5)  
    btn_1 = Button(GG, text="No", fg='white',bg='#E64345', height = 5, width = 5, command=fermer)
    btn_1.pack(padx=5, pady=5)  
    #newWindow.iconbitmap('Back Up\\icon\\info.ico')

def Score(temps,niveau):

    gui = tk.Tk()
    gui.title("Your name")
    gui.geometry("300x100")

    def getEntry(temps,gui,niveau):
        name = myEntry.get()
        gui.destroy()
        afficher(name,temps)
        if niveau=="easy":
            enregisterE(name,temps)
        if niveau=="normal":
            enregisterN(name,temps)
        if niveau=="hard":
            enregisterH(name,temps)  
        
        
    myEntry = tk.Entry(gui, width=40)
    myEntry.pack(pady=20)
    btn = tk.Button(gui, height=1, width=10, text="Your name", command=lambda: [getEntry(temps,gui,niveau)])
    btn.pack()

def play(niveau):

    tic = time.perf_counter()
    if niveau=="easy":
        number=random.randint(1, 999)
    if niveau=="normal":
        number=random.randint(1, 9999)
    if niveau=="hard":
        number=round(random.uniform(0.01, 9999.99),2)

    gui = tk.Tk()
    gui.title("Find the number")
    gui.geometry("300x100")

    def verification(number,niveau):
        res = float(myEntry.get())
        float(res)
        if number==res:
            print("Gagnée ")
            toc = time.perf_counter()
            temps=(toc-tic)
            gui.destroy()
            Score(temps,niveau)

        if number>res:
            print("plus grand ")
        if number<res:
            print("plus petit ")


    myEntry = tk.Entry(gui, width=40)
    myEntry.pack(pady=20)
    btn = tk.Button(gui, height=1, width=10, text="Lire", command=lambda: [verification(number,niveau)])
    btn.pack()

# Information page
def new_window():

    # be treated as a new window
    newWindow = Toplevel(fenetre)
      
    # set the title 
    newWindow.title("Informations !")
      
    # sets the dimensions
    newWindow.geometry("280x260")
      
    # in the Window
    Label(newWindow, 
        text ="\n" + "Welcome to the information page" + "\n",fg='#1E90FF',font='bold').pack()
    Label(newWindow, 
        text ="[Classement] : Allows you to see the scoreboard" + "\n",fg='orange').pack()
    Label(newWindow, 
        text ="[Easy] : Set the difficulty on easy",fg='dark blue').pack()
    Label(newWindow, 
        text ="[Normal] : Set the difficulty on normal",fg='dark blue').pack()
    Label(newWindow, 
        text ="[Hard] : Set the difficulty on hard",fg='dark blue').pack()
    Label(newWindow, 
        text ="\n" + "Thanks !" + "\n",fg='#A965CA',font='bold').pack()
    btn_5 = Button(newWindow, text="Exit", fg ='white',bg='black', command=newWindow.destroy)
    btn_5.pack(padx=5, pady=5)
    #newWindow.iconbitmap('Back Up\\icon\\info.ico')

def main():
    fenetre = Tk()
    # set the title of the main window
    fenetre.title("PYCode")
    # set the dimension of the main window
    fenetre.geometry("450x125")
    Label(fenetre, text ="\n" + "Welcome !" + "\n" "Are you ready to play ?" + "\n",fg='Black',font='bold').pack()

    btn_0 = Button(fenetre, text="Hard", fg='white',bg='#ff2626', height = 5, width = 8, command=Hard)
    btn_0.pack(side=RIGHT, padx=5, pady=5)
    btn_1 = Button(fenetre, text="Normal", fg='white',bg='#9ADBC5', height = 5, width = 8, command=Normal)
    btn_1.pack(side=RIGHT, padx=5, pady=5)  
    btn_2 = Button(fenetre, text="Easy", fg='white',bg='#A0DDE0', height = 5, width = 8, command=Easy)
    btn_2.pack(side=RIGHT, padx=5, pady=5)  
    btn_3 = Button(fenetre, text="Classement", fg='white',bg='#FDC453', height = 5, width = 10, command=Classement)
    btn_3.pack(side=LEFT, padx=5, pady=5) 
    btn_4 = Button(fenetre, text="Info", fg='white',bg='#3D4048', height = 5, width = 6, command=new_window)
    btn_4.pack(side=LEFT, padx=5, pady=5)  

    #fenetre.iconbitmap('Back Up\\icon\\icon.ico')
    fenetre.mainloop()

if __name__ == "__main__":
    main()
